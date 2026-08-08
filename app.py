import plotly.express as px
import plotly
import json
from config import *
import csv
import logging
import random
from zoneinfo import ZoneInfo
from datetime import datetime, timedelta
from flask import session
from email_sender import send_otp
from datetime import datetime, timedelta
from flask import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from flask import flash
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import io
from reportlab.graphics.shapes import Drawing, Line
from flask import Flask, render_template, request, redirect, session
from flask import Flask, render_template, request, redirect
from werkzeug.security import generate_password_hash
from werkzeug.security import check_password_hash
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image
)

from reportlab.lib import colors

# ==========================================
# Permanent Super Admin
# ==========================================

SUPER_ADMIN_USERNAME = "@Aviroy.2207"

from reportlab.lib.styles import getSampleStyleSheet

from reportlab.lib.enums import TA_CENTER
from flask import Response, send_file
import csv
import io
from predict import predict_sentiment
from pdf_report import generate
from flask import send_file
from suggestion import get_suggestion
from database import get_connection
from datetime import datetime
current_time = datetime.now().strftime("%d %B %Y | %I:%M %p")

# ==========================================
# Indian Standard Time Helper
# ==========================================

def get_ist_time():
    return datetime.now(
        ZoneInfo("Asia/Kolkata")
    )

def get_ist_string():
    return get_ist_time().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

# ==========================================
# Notification Helper
# ==========================================

def add_notification(message, notification_type):

    conn = get_connection()

    cursor = conn.cursor()

    cursor.execute("""
    INSERT INTO notifications(

        message,

        type,

        created_at

    )

    VALUES(?,?,?)

    """, (

        message,

        notification_type,

        get_ist_string()

    ))

    conn.commit()

    conn.close()

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY")
app.secret_key = "smart_feedback_portal_2026"
from datetime import timedelta

# ==========================================
# Session Configuration
# ==========================================

app.permanent_session_lifetime = timedelta(minutes=30)

# ==========================================
# Auto Refresh Session Timer
# ==========================================

@app.before_request
def refresh_session():

    session.permanent = True

    session.modified = True

# ==========================================
# Session Expiry Check
# ==========================================

from flask import request

@app.before_request
def check_session():

    public_routes = [
        "home",
        "login",
        "register",
        "forgot_password",
        "verify_register_otp",
        "verify_forgot_otp",
        "admin_login",
        "verify_otp",
        "resend_otp",
        "resend_register_otp",
        "static"
    ]

    if request.endpoint in public_routes:
        return

    # -------------------------------
    # Student Protected Pages
    # -------------------------------

    student_pages = [
        "/dashboard",
        "/feedback",
        "/submit_feedback",
        "/profile"
    ]

    if request.path in student_pages:

        if "user" not in session:

            flash(
                "⏰ Your session expired. Please login again.",
                "warning"
            )

            return redirect("/login")

    # -------------------------------
    # Admin Protected Pages
    # -------------------------------

    admin_pages = [

        "/dashboard_v2",

        "/create_admin",

        "/manage_admin",

        "/student_management",

        "/ai_assistant"

    ]

    # Pages with URL parameters
    if request.path.startswith("/view_student"):
        admin_required = True
    elif request.path.startswith("/edit_admin"):
        admin_required = True
    elif request.path.startswith("/delete_admin"):
        admin_required = True
    elif request.path.startswith("/delete_student"):
        admin_required = True
    else:
        admin_required = request.path in admin_pages

    if admin_required:

        if "admin" not in session:

            flash(
                "⏰ Admin session expired. Please login again.",
                "warning"
            )

            return redirect("/admin_login")

app.secret_key=SECRET_KEY

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/register", methods=["GET", "POST"])
def register():

    if request.method == "POST":

        name = request.form["name"]
        email = request.form["email"]
        password = request.form["password"]

        conn = get_connection()
        cursor = conn.cursor()

        # Check whether email already exists
        cursor.execute(
            "SELECT * FROM students WHERE email=?",
            (email,)
        )

        existing_user = cursor.fetchone()

        conn.close()

        if existing_user:

            flash(
                "Email already registered!",
                "warning"
            )

            return redirect("/register")

        # Generate OTP
        otp = str(random.randint(100000, 999999))

        # Save registration data temporarily
        session["student_name"] = name
        session["student_email"] = email
        session["student_password"] = password
        session["student_otp"] = otp
        session["student_otp_time"] = datetime.now().isoformat()

        # Send OTP
        if send_otp(email, otp):

            flash(
                "OTP has been sent to your email.",
                "success"
            )

            return redirect("/verify_register_otp")

        else:

            flash(
                "Unable to send OTP. Please check your email address and try again.",
                "danger"
            )

            return redirect("/register")

    return render_template("register.html")

@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        email = request.form["email"]
        password = request.form["password"]

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT *
        FROM students
        WHERE email=?
        """, (email,))

        user = cursor.fetchone()

        if user and check_password_hash(
            user["password"],
            password
        ):

            # Update student status
            cursor.execute("""
            UPDATE students
            SET
                last_login=?,
                status='Online'
            WHERE email=?
            """, (
                get_ist_string(),
                email
            ))

            # Save login history
            cursor.execute("""
            INSERT INTO login_history(

                student_name,

                email,

                login_time,

                status

            )

            VALUES(?,?,?,?)
            """, (

                user["name"],

                user["email"],

                get_ist_string(),

                "Online"

            ))

            conn.commit()

            session.permanent = True
            session["user"] = user["name"]
            session["email"] = user["email"]

            logging.info(
                f"Student Logged In: {user['name']}"
            )

            conn.close()

            return redirect("/dashboard")

        conn.close()

        flash(
            "Invalid Email or Password",
            "danger"
        )

        return redirect("/login")

    return render_template("login.html")

@app.route("/forgot_password", methods=["GET", "POST"])
def forgot_password():

    # -----------------------------
    # STEP 1 : Send OTP
    # -----------------------------
    if request.method == "POST" and not session.get("otp_verified"):

        email = request.form["email"]

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT * FROM students
        WHERE email=?
        """, (email,))

        user = cursor.fetchone()

        conn.close()

        if user:

            otp = str(random.randint(100000,999999))

            session["reset_email"] = email
            session["reset_otp"] = otp
            session["reset_otp_time"] = datetime.now().isoformat()

            send_otp(email, otp)

            flash(
                "OTP sent successfully.",
                "success"
            )

            return redirect("/verify_reset_otp")

        else:

            flash(
                "Email not found!",
                "danger"
            )

            return redirect("/forgot_password")

    # -----------------------------
    # STEP 2 : Update Password
    # -----------------------------
    if request.method == "POST" and session.get("otp_verified"):

        password = request.form["password"]

        confirm = request.form["confirm_password"]

        if password != confirm:

            flash(
                "Passwords do not match!",
                "danger"
            )

            return render_template(
                "forgot_password.html",
                otp_verified=True
            )

        hashed_password = generate_password_hash(password)

        conn = get_connection()

        cursor = conn.cursor()

        cursor.execute("""
        UPDATE students
        SET password=?
        WHERE email=?
        """, (
            hashed_password,
            session["reset_email"]
        ))

        conn.commit()

        session.pop("otp_verified", None)
        session.pop("reset_email", None)
        session.pop("reset_otp", None)
        session.pop("reset_otp_time", None)

        conn.close()

        # Clear OTP session
        session.pop("reset_email", None)
        session.pop("reset_otp", None)
        session.pop("reset_otp_time", None)
        session.pop("otp_verified", None)

        flash(
            "✅ Password Updated Successfully!",
            "success"
        )

        return redirect("/login")

    print("otp_verified =", session.get("otp_verified"))
    print(dict(session))

    return render_template(
        "forgot_password.html",
        otp_verified=session.get("otp_verified", False)
    )

@app.route("/admin_forgot_password", methods=["GET", "POST"])
def admin_forgot_password():

    if request.method == "POST":

        username = request.form["username"].strip()
        email = request.form["email"].strip()

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT *
        FROM admin
        WHERE username=? AND email=?
        """, (username, email))

        admin = cursor.fetchone()

        conn.close()

        if admin:

            otp = str(random.randint(100000, 999999))

            session["admin_reset_username"] = username
            session["admin_reset_email"] = email
            session["admin_reset_otp"] = otp
            session["admin_reset_otp_time"] = datetime.now().isoformat()

            send_otp(email, otp)

            flash(
                "OTP sent successfully.",
                "success"
            )

            return redirect("/admin_verify_reset_otp")

        flash(
            "Invalid Username or Email!",
            "danger"
        )

        return redirect("/admin_forgot_password")

    return render_template("admin_forgot_password.html")

@app.route("/dashboard")
def dashboard():

    if "user" not in session:
        return redirect("/login")

    return render_template(
        "dashboard.html",
        username=session["user"]
    )

@app.route("/feedback", methods=["GET", "POST"])
def feedback():

    if "user" not in session:
        return redirect("/login")

    if request.method == "POST":

        conn = get_connection()
        cursor = conn.cursor()

        # ============================================
        # Generate Next Submission ID
        # ============================================

        cursor.execute("""
        SELECT MAX(submission_id)
        FROM feedback
        """)

        last_submission = cursor.fetchone()[0]

        if last_submission is None:
            submission_id = 1
        else:
            submission_id = last_submission + 1

        categories = [

            (
                "Course",
                request.form["course_rating"],
                request.form.get("course_comment", "")
            ),

            (
                "Faculty",
                request.form["faculty_rating"],
                request.form.get("faculty_comment", "")
            ),

            (
                "Facility",
                request.form["facility_rating"],
                request.form.get("facility_comment", "")
            )
            

        ]

        overall_suggestion = request.form.get(
            "overall_suggestion",
            ""
        ).strip()

        for category, rating, comment in categories:

            # Combine rating + comment
            feedback_text = f"Rating: {rating}"

            if comment.strip():

                feedback_text += f"\n\nComment:\n{comment.strip()}"

            sentiment = predict_sentiment(feedback_text)

            suggestion = get_suggestion(sentiment)

            cursor.execute("""

            INSERT INTO feedback
            (
            submission_id,
            student_name,
            category,
            feedback,
            sentiment,
            suggestion,
            overall_suggestion,
            date
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)

            """,
            (
            submission_id,
            session["user"],
            category,
            feedback_text,
            sentiment,
            suggestion,
            overall_suggestion,
            get_ist_string()
            )
            )

        conn.commit()

        conn.close()

        logging.info(
            f"{session['user']} submitted complete feedback."
        )

        flash(
            "✅ Complete feedback submitted successfully!",
            "success"
        )

        return redirect("/view_feedback")

    return render_template("feedback.html")

@app.route("/view_feedback")
def view_feedback():

    if "user" not in session:
        return redirect("/login")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM feedback
        WHERE student_name=?
        ORDER BY submission_id DESC,
        CASE category
            WHEN 'Course' THEN 1
            WHEN 'Faculty' THEN 2
            WHEN 'Facility' THEN 3
        END
    """, (session["user"],))

    rows = cursor.fetchall()
    conn.close()

    submissions = []
    current_submission = None

    for row in rows:

        if (current_submission is None or
                current_submission["submission_id"] != row["submission_id"]):

            current_submission = {

                "submission_id": row["submission_id"],

                "date": row["date"],

                "overall_suggestion": row["overall_suggestion"],

                "items": []

            }

            submissions.append(current_submission)

        current_submission["items"].append(row)

    return render_template(

        "view_feedback.html",

        submissions=submissions

    )

@app.route("/delete_feedback/<int:id>")
def delete_feedback(id):

    conn = get_connection()

    cursor = conn.cursor()

    cursor.execute(
        """
        DELETE FROM feedback
        WHERE id=?
        AND student_name=?
        """,
        (
            id,
            session["user"]
        )
    )

    conn.commit()

    conn.close()

    return redirect("/view_feedback?deleted=1")

@app.route("/logout")
def logout():

    if "user" in session:

        conn = get_connection()
        cursor = conn.cursor()

        # Student Offline
        cursor.execute("""
        UPDATE students
        SET status='Offline'
        WHERE email=?
        """, (
            session["email"],
        ))

        # Update latest login history
        cursor.execute("""
        UPDATE login_history

        SET

        logout_time=?,

        status='Logged Out'

        WHERE id=(

            SELECT id

            FROM login_history

            WHERE email=?

            ORDER BY id DESC

            LIMIT 1

        )
        """, (

            get_ist_string(),

            session["email"]

        ))

        conn.commit()
        conn.close()

    session.clear()

    return redirect("/")

from werkzeug.security import check_password_hash


@app.route("/admin_login", methods=["GET", "POST"])
def admin_login():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        print("USERNAME =", repr(username))
        print("PASSWORD =", repr(password))

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT *
        FROM admin
        WHERE username=?
        """, (username,))

        admin = cursor.fetchone()

        conn.close()

        if admin and check_password_hash(
            admin["password"],
            password
        ):

            session.permanent = True
            session["admin"] = admin["username"]

            logging.info(
                f"Admin Logged In: {admin['username']}"
            )

            add_notification(
                f"🟢 Administrator '{admin['username']}' logged into the portal successfully.",
                "admin"
            )

            return redirect("/dashboard_v2")

        flash(
            "Invalid Username or Password!",
            "danger"
        )

        return redirect("/admin_login")

    return render_template("admin_login.html")

@app.route("/admin_dashboard")
def admin_dashboard():

    if "admin" not in session:
        return redirect("/admin_login")

    conn = get_connection()
    cursor = conn.cursor()

    # ----------------------------
    # Feedback Table
    # ----------------------------

    cursor.execute("""
        SELECT *
        FROM feedback
        ORDER BY id DESC
    """)

    feedback = cursor.fetchall()

    # ----------------------------
    # Dashboard Cards
    # ----------------------------

    cursor.execute("SELECT COUNT(*) FROM feedback")
    total = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM feedback
        WHERE sentiment='Positive'
    """)
    positive = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM feedback
        WHERE sentiment='Negative'
    """)
    negative = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM feedback
        WHERE sentiment='Neutral'
    """)
    neutral = cursor.fetchone()[0]

    # ----------------------------
    # Category Count
    # ----------------------------

    cursor.execute("""
        SELECT category,
               COUNT(*) AS total
        FROM feedback
        GROUP BY category
        ORDER BY category
    """)

    category_data = cursor.fetchall()

    conn.close()

    # =====================================
    # PIE CHART
    # =====================================

    pie = px.pie(

        names=[
            "Positive",
            "Negative",
            "Neutral"
        ],

        values=[
            positive,
            negative,
            neutral
        ],

        hole=.45,

        title="Feedback Sentiment Analysis",

        color=[
            "Positive",
            "Negative",
            "Neutral"
        ],

        color_discrete_map={

            "Positive":"#198754",

            "Negative":"#dc3545",

            "Neutral":"#ffc107"

        }

    )

    pie.update_traces(

        textinfo="label+percent",

        textfont_size=15,

        hovertemplate="<b>%{label}</b><br>Count : %{value}<br>%{percent}<extra></extra>"

    )

    pie.update_layout(

        title_x=.5,

        template="plotly_white",

        height=420,

        margin=dict(l=20,r=20,t=60,b=20),

        legend_title="Sentiment"

    )

    # =====================================
    # BAR CHART
    # =====================================

    category=[]

    count=[]

    for row in category_data:

        category.append(row["category"])

        count.append(row["total"])

    bar = px.bar(

        x=category,

        y=count,

        text=count,

        color=category,

        title="Feedback Category Analysis"

    )

    bar.update_traces(

        textposition="outside"

    )

    bar.update_layout(

        template="plotly_white",

        title_x=.5,

        height=420,

        showlegend=False,

        xaxis_title="Category",

        yaxis_title="Number of Feedback",

        yaxis=dict(dtick=1)

    )

    return render_template(

        "admin_dashboard.html",

        feedback=feedback,

        total=total,

        positive=positive,

        negative=negative,

        neutral=neutral,

        pie_graph=pie.to_json(),

        bar_graph=bar.to_json()

    )

@app.route("/admin_delete/<int:id>")
def admin_delete(id):

    conn=get_connection()

    cursor=conn.cursor()

    cursor.execute(

        "DELETE FROM feedback WHERE id=?",

        (id,)

    )

    conn.commit()

    conn.close()

    flash(
        "Feedback deleted successfully.",
        "success"
    )

    return redirect("/dashboard_v2")

@app.route("/admin_logout")
def admin_logout():

    session.pop("admin",None)

    return redirect("/")

import io
import csv
from flask import Response

@app.route("/export")
def export():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT *
    FROM feedback
    ORDER BY submission_id DESC,
    CASE category
        WHEN 'Course' THEN 1
        WHEN 'Faculty' THEN 2
        WHEN 'Facility' THEN 3
    END
    """)

    rows = cursor.fetchall()

    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "S.No.",
        "Submission ID",
        "Student",
        "Category",
        "Feedback",
        "Sentiment",
        "Suggestion",
        "Date"
    ])

    for i, row in enumerate(rows, start=1):

        writer.writerow([
            i,
            row["submission_id"],
            row["student_name"],
            row["category"],
            row["feedback"],
            row["sentiment"],
            row["suggestion"],
            row["date"]
        ])

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition":
            "attachment; filename=feedback.csv"
        }
    )

@app.route("/export_excel")
def export_excel():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT *
    FROM feedback
    ORDER BY submission_id DESC,
    CASE category
        WHEN 'Course' THEN 1
        WHEN 'Faculty' THEN 2
        WHEN 'Facility' THEN 3
    END
    """)

    rows = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback Report"

    # ==========================
    # Report Title
    # ==========================

    ws.merge_cells("A1:I1")

    title = ws["A1"]
    title.value = "SMART FEEDBACK & SUGGESTION PORTAL"
    ws.merge_cells("A2:I2")
    ws["A2"] = "Complete Feedback Report"
    ws["A2"].font = Font(size=14, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    title.font = Font(size=16, bold=True, color="0D6EFD")
    title.alignment = Alignment(horizontal="center")

    # ==========================
    # Header
    # ==========================

    headers = [

        "S.No.",
        "Submission ID",
        "Student Name",
        "Category",
        "Feedback",
        "Sentiment",
        "AI Suggestion",
        "Student Suggestion",
        "Date"

    ]

    header_fill = PatternFill(
        start_color="0D6EFD",
        end_color="0D6EFD",
        fill_type="solid"
    )

    thin = Side(style="thin")

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(row=2, column=col)

        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

    # ==========================
    # Data
    # ==========================

    previous_submission = None

    excel_row = 3

    for index, row in enumerate(rows, start=1):

        ws.cell(excel_row, 1).value = index
        ws.cell(excel_row, 2).value = row["submission_id"]
        ws.cell(excel_row, 3).value = row["student_name"]
        ws.cell(excel_row, 4).value = row["category"]

        feedback_cell = ws.cell(excel_row, 5)
        feedback_cell.value = row["feedback"]
        feedback_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top"
        )

        ws.cell(excel_row, 6).value = row["sentiment"]
        ws.cell(excel_row, 7).value = row["suggestion"]

        if previous_submission != row["submission_id"]:
            ws.cell(excel_row, 8).value = row["overall_suggestion"]
        else:
            ws.cell(excel_row, 8).value = ""

        ws.cell(excel_row, 9).value = row["date"]

        # Borders
        for col in range(1, 10):
            ws.cell(excel_row, col).border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

        previous_submission = row["submission_id"]
        excel_row += 1

    # ==========================
    # Freeze Header
    # ==========================

    ws.freeze_panes = "A3"

    # ==========================
    # Filters
    # ==========================

    ws.auto_filter.ref = f"A2:I{excel_row-1}"

    # ==========================
    # Auto Column Width
    # ==========================

    from openpyxl.utils import get_column_letter

    for col in range(1, 10):

        max_length = 0

        column_letter = get_column_letter(col)

        for row in range(2, excel_row):

            cell = ws.cell(row=row, column=col)

            if cell.value:

                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 5, 50)

    filename = "Smart_Feedback_Report.xlsx"

    wb.save(filename)

    return send_file(
        filename,
        as_attachment=True
    )


@app.route("/pdf")
def pdf():

    generate()

    return send_file(
        "FeedbackReport.pdf",
        as_attachment=True
    )

@app.errorhandler(404)

def page_not_found(error):

    return render_template(
        "404.html"
    ),404

@app.route("/dashboard_v2")
def dashboard_v2():

    conn = get_connection()
    cursor = conn.cursor()

    # Dashboard Cards
    cursor.execute("SELECT COUNT(*) FROM feedback")
    total = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM feedback WHERE sentiment='Positive'")
    positive = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM feedback WHERE sentiment='Negative'")
    negative = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM feedback WHERE sentiment='Neutral'")
    neutral = cursor.fetchone()[0]

    # ==========================
    # Total Students
    # ==========================

    cursor.execute("""
    SELECT COUNT(*)
    FROM students
    """)

    total_students = cursor.fetchone()[0]


    # ==========================
    # Online Students
    # ==========================

    cursor.execute("""
    SELECT COUNT(*)
    FROM students
    WHERE status='Online'
    """)

    online_students = cursor.fetchone()[0]


    # ==========================
    # Today's Student Registrations
    # ==========================

    cursor.execute("""
    SELECT COUNT(*)
    FROM students
    WHERE DATE(register_date)=DATE(?)
    """, (
        get_ist_string()[:10],
    ))

    today_students = cursor.fetchone()[0]

    # Category Count
    cursor.execute("""
    SELECT category, COUNT(*)
    FROM feedback
    GROUP BY category
    """)

    rows = cursor.fetchall()

    cursor.execute("""
    SELECT
    DATE(date) as feedback_date,
    COUNT(*) as total
    FROM feedback
    GROUP BY DATE(date)
    ORDER BY DATE(date)
    """)
    trend_rows = cursor.fetchall()

    trend_dates = []
    trend_counts = []

    for row in trend_rows:

        trend_dates.append(row["feedback_date"])

        trend_counts.append(row["total"])

    bar_category = []
    bar_count = []

    for row in rows:
        bar_category.append(row["category"])
        bar_count.append(row[1])

    cursor.execute("""
    SELECT student_name,
    category,
    sentiment,
    date
    FROM feedback
    ORDER BY id DESC
    LIMIT 5
    """)
    recent_feedback = cursor.fetchall()

    cursor.execute("""
    SELECT *
    FROM feedback
    ORDER BY submission_id DESC,
    CASE category
        WHEN 'Course' THEN 1
        WHEN 'Faculty' THEN 2
        WHEN 'Facility' THEN 3
    END
    """)

    feedback = cursor.fetchall()

    # ==========================
    # Today's Feedback
    # ==========================

    cursor.execute("""
    SELECT COUNT(*)
    FROM feedback
    WHERE DATE(date)=DATE('now')
    """)

    today_feedback = cursor.fetchone()[0]


    # ==========================
    # This Month Feedback
    # ==========================

    cursor.execute("""
    SELECT COUNT(*)
    FROM feedback
    WHERE DATE(date) = CURRENT_DATE
    """)

    month_feedback = cursor.fetchone()[0]

    conn.close()

    # Pie Chart
    pie_labels = ["Positive", "Negative", "Neutral"]
    pie_values = [positive, negative, neutral]

    # Percentages
    positive_percent = round((positive / total) * 100, 1) if total else 0
    negative_percent = round((negative / total) * 100, 1) if total else 0
    neutral_percent = round((neutral / total) * 100, 1) if total else 0

    # Most Discussed Category
    if bar_count:

        max_index = bar_count.index(max(bar_count))
        top_category = bar_category[max_index]

    else:

        top_category = "No Feedback"
    

    satisfaction_score = round(
        ((positive * 100) + (neutral * 50)) / total
        ) if total else 0
    

    # ==========================
    # AI Health Score
    # ==========================

    health_score = satisfaction_score

    if total == 0:

        health_score = None
        health_status = "No Data"
        health_color = "secondary"

    else:

        health_score = satisfaction_score

        if health_score >= 80:

            health_status = "Excellent"
            health_color = "success"

        elif health_score >= 60:

            health_status = "Good"
            health_color = "primary"

        elif health_score >= 40:

            health_status = "Needs Improvement"
            health_color = "warning"

        else:

            health_status = "Critical"
            health_color = "danger"
    
    if total == 0:

        satisfaction_stars = "No Rating"

    elif satisfaction_score >= 80:

        satisfaction_stars = "⭐⭐⭐⭐⭐"

    elif satisfaction_score >= 60:

        satisfaction_stars = "⭐⭐⭐⭐"

    elif satisfaction_score >= 40:

        satisfaction_stars = "⭐⭐⭐"

    elif satisfaction_score >= 20:

        satisfaction_stars = "⭐⭐"

    else:

        satisfaction_stars = "⭐"


    # =====================================
    # Dynamic AI Recommendation Engine
    # =====================================

    recommendation = []

    # Overall satisfaction
    if total == 0:

        recommendation.append(
            "No feedback has been submitted yet. AI recommendations will appear after students submit feedback."
        )

    elif positive_percent >= 70:

        recommendation.append(
            "Student satisfaction is excellent. Maintain the current academic standards."
        )

    elif positive_percent >= 50:

        recommendation.append(
            "Overall feedback is positive, but there is still room for improvement."
        )

    else:

        recommendation.append(
            "Student satisfaction is low. Immediate attention is recommended."
        )


    # Negative feedback
    if negative_percent >= 40:

        recommendation.append(
            "Negative feedback is high. Review student complaints carefully."
        )


    # Category-based recommendations
    if top_category == "Course":

        recommendation.append(
            "Review the course syllabus and teaching methodology."
        )

    elif top_category == "Faculty":

        recommendation.append(
            "Conduct faculty development and communication workshops."
        )

    elif top_category == "Facility":

        recommendation.append(
            "Improve laboratory facilities and campus infrastructure."
        )

    elif top_category == "No Feedback":

        recommendation.append(
            "No feedback has been submitted yet."
        )

    # Neutral feedback
    if neutral_percent >= 20:

        recommendation.append(
            "Encourage students to provide more detailed feedback."
        )


    # Health score
    if total > 0 and health_score < 50:

        recommendation.append(
            "Priority action is required to improve student satisfaction."
        )

    # AI Prediction
    if total == 0:

        prediction = "No feedback available for AI prediction."

    elif positive_percent >= 70:

        prediction = "Student satisfaction is excellent."

    elif positive_percent >= 50:

        prediction = "Overall feedback is positive, but there is room for improvement."

    else:

        prediction = "Immediate improvements are recommended based on student feedback."

    # ==========================
    # Unread Notifications
    # ==========================

    cursor = get_connection().cursor()

    cursor.execute("""
    SELECT COUNT(*)
    FROM notifications
    WHERE is_read = 0
    """)

    notification_count = cursor.fetchone()[0]

    # ==========================
    # System Status
    # ==========================

    database_status = "Connected"
    database_color = "success"

    ai_status = "Active"
    ai_color = "success"

    notification_status = "Running"
    notification_color = "success"

    timezone = "Asia/Kolkata (IST)"

    session_timeout = "30 Minutes"

    return render_template(
        "dashboard_v2.html",
        total=total,
        positive=positive,
        negative=negative,
        neutral=neutral,
        pie_labels=pie_labels,
        pie_values=pie_values,
        bar_category=bar_category,
        bar_count=bar_count,
        positive_percent=positive_percent,
        negative_percent=negative_percent,
        neutral_percent=neutral_percent,
        top_category=top_category,
        recommendation=recommendation,
        satisfaction_score=satisfaction_score,
        satisfaction_stars=satisfaction_stars,
        prediction=prediction,
        recent_feedback=recent_feedback,
        feedback=feedback,
        current_time=current_time,
        trend_dates=trend_dates,
        trend_counts=trend_counts,
        today_feedback=today_feedback,
        month_feedback=month_feedback,
        health_score=health_score,
        health_status=health_status,
        health_color=health_color,
        notification_count=notification_count,
        database_status=database_status,
        database_color=database_color,

        ai_status=ai_status,
        ai_color=ai_color,

        notification_status=notification_status,
        notification_color=notification_color,

        timezone=timezone,

        session_timeout=session_timeout,
        total_students=total_students,
        online_students=online_students,
        today_students=today_students,
        )

@app.route("/create_admin", methods=["GET", "POST"])
def create_admin():

    if request.method == "POST":

        name = request.form["name"]
        email = request.form["email"]
        username = request.form["username"]
        password = request.form["password"]

        # Generate a 6-digit OTP
        otp = str(random.randint(100000, 999999))



        # Save data temporarily in the session
        session["otp"] = otp
        session["otp_time"] = datetime.now().isoformat()
        session["name"] = name
        session["email"] = email
        session["username"] = username
        session["password"] = password

        # Send the OTP email
        send_otp(email, otp)

        # Go to the OTP verification page
        return redirect("/verify_otp")

        conn = get_connection()
        cursor = conn.cursor()

        # Check if username already exists
        cursor.execute(
            "SELECT * FROM admin WHERE username=?",
            (username,)
        )

        existing_admin = cursor.fetchone()

        if existing_admin:

            conn.close()

            return render_template(
                "create_admin.html",
                error="❌ Username already exists!"
            )

        # Insert new admin
        cursor.execute("""
        INSERT INTO admin(username, password)
        VALUES (?, ?)
        """, (username, password))

        conn.commit()
        conn.close()

        return render_template(
            "create_admin.html",
            success="✅ New Administrator Created Successfully!"
        )

    # If the request is GET, just show the page
    return render_template("create_admin.html")

@app.route("/admin_management")
def admin_management():

    conn = get_connection()

    cursor = conn.cursor()

    # Fetch all admins
    cursor.execute("SELECT * FROM admin")
    admins = cursor.fetchall()

    # Total Admins

    cursor.execute("SELECT COUNT(*) FROM admin")

    total_admins=cursor.fetchone()[0]

    # Protected Admin

    protected_admins=1

    # Active Admins

    active_admins=total_admins-protected_admins

    conn.close()

    return render_template(

        "admin_management.html",

        admins=admins,

        total_admins=total_admins,

        active_admins=active_admins,

        protected_admins=protected_admins,
        
        super_admin=SUPER_ADMIN_USERNAME

        )


@app.route("/delete_admin/<int:id>")
def delete_admin(id):

    conn = get_connection()
    cursor = conn.cursor()

    # Get admin details
    cursor.execute("""
    SELECT *
    FROM admin
    WHERE id=?
    """, (id,))

    admin = cursor.fetchone()

    # Prevent deleting Super Admin
    if admin and admin["username"] == SUPER_ADMIN_USERNAME:

        conn.close()

        flash(
            "❌ Super Admin cannot be deleted.",
            "danger"
        )

        return redirect("/admin_management")

    cursor.execute(
        "DELETE FROM admin WHERE id=?",
        (id,)
    )

    conn.commit()
    conn.close()

    return redirect("/admin_management?deleted=1")

@app.route("/edit_admin/<int:id>", methods=["GET", "POST"])
def edit_admin(id):

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT *
    FROM admin
    WHERE id=?
    """, (id,))

    admin = cursor.fetchone()

    # Prevent editing Super Admin
    if admin and admin["username"] == SUPER_ADMIN_USERNAME:

        conn.close()

        flash(
            "❌ Super Admin cannot be edited.",
            "danger"
        )

        return redirect("/admin_management")

    if request.method == "POST":

        username = request.form["username"].strip()
        password = request.form["password"].strip()

        # ==========================================
        # Check if username already exists
        # ==========================================

        cursor.execute("""
        SELECT id
        FROM admin
        WHERE username=?
        AND id!=?
        """, (username, id))

        existing = cursor.fetchone()

        if existing:

            conn.close()

            flash(
                "Username already exists.",
                "danger"
            )

            return redirect(f"/edit_admin/{id}")

        # Update username only
        if password == "":

            cursor.execute("""
            UPDATE admin
            SET username=?
            WHERE id=?
            """, (
                username,
                id
            ))

        # Update username + password
        else:

            cursor.execute("""
            UPDATE admin
            SET username=?,
                password=?
            WHERE id=?
            """, (
                username,
                generate_password_hash(password),
                id
            ))

        conn.commit()
        conn.close()

        return redirect("/admin_management?updated=1")

    conn.close()

    return render_template(
        "edit_admin.html",
        admin=admin
    )

@app.route("/test_email")
def test_email():

    # Replace with YOUR OWN email address
    receiver = "smartfeedbackportal@gmail.com"

    send_otp(receiver, "123456")

    return "✅ Email Sent Successfully!"

@app.route("/verify_otp", methods=["GET", "POST"])
def verify_otp():

    if request.method == "POST":

        user_otp = request.form["otp"]

        # ===============================
        # Check OTP Expiry (5 Minutes)
        # ===============================
        if "otp_time" in session:

            otp_time = datetime.fromisoformat(session["otp_time"])

            if datetime.now() > otp_time + timedelta(minutes=5):

                session.clear()

                return render_template(
                    "verify_otp.html",
                    error="❌ OTP Expired! Please request a new OTP.",
                    remaining=0
                )

        # ===============================
        # Verify OTP
        # ===============================
        if user_otp == session.get("otp"):

            conn = get_connection()
            cursor = conn.cursor()

            # Check duplicate username
            cursor.execute(
                "SELECT * FROM admin WHERE username=?",
                (session["username"],)
            )

            if cursor.fetchone():

                conn.close()

                return render_template(
                    "verify_otp.html",
                    error="❌ Username already exists!",
                    remaining=0
                )

            # Check duplicate email
            cursor.execute(
                "SELECT * FROM admin WHERE email=?",
                (session["email"],)
            )

            if cursor.fetchone():

                conn.close()

                return render_template(
                    "verify_otp.html",
                    error="❌ Email already exists!",
                    remaining=0
                )

            # ===============================
            # Create Admin
            # ===============================
            cursor.execute("""
                INSERT INTO admin(name,email,username,password)
                VALUES(?,?,?,?)
            """, (
                session["name"],
                session["email"],
                session["username"],
                generate_password_hash(session["password"])
            ))

            conn.commit()
            conn.close()

            # Clear session
            session.clear()

            return render_template(
                "create_admin.html",
                success="✅ Administrator Created Successfully!"
            )

        else:

            # Remaining resend time
            remaining = 0

            if "resend_time" in session:

                last_time = datetime.fromisoformat(session["resend_time"])

                remaining = 30 - (datetime.now() - last_time).seconds

                if remaining < 0:

                    remaining = 0

            return render_template(
                "verify_otp.html",
                error="❌ Invalid OTP!",
                remaining=remaining
            )

    # ======================================
    # GET Request
    # ======================================

    remaining = 0

    if "resend_time" in session:

        last_time = datetime.fromisoformat(session["resend_time"])

        remaining = 30 - (datetime.now() - last_time).seconds

        if remaining < 0:

            remaining = 0

    return render_template(
        "verify_otp.html",
        remaining=remaining
    )


@app.route("/admin_verify_reset_otp", methods=["GET", "POST"])
def admin_verify_reset_otp():

    if request.method == "POST":

        user_otp = request.form["otp"]

        otp_time = datetime.fromisoformat(
            session["admin_reset_otp_time"]
        )

        if datetime.now() > otp_time + timedelta(minutes=5):

            session.pop("admin_reset_otp", None)

            flash(
                "❌ OTP Expired! Please request a new OTP.",
                "danger"
            )

            return redirect("/admin_forgot_password")

        if user_otp == session.get("admin_reset_otp"):

            session["admin_otp_verified"] = True

            return redirect("/admin_reset_password")

        flash(
            "❌ Invalid OTP!",
            "danger"
        )

    return render_template("verify_reset_otp.html")

@app.route("/admin_reset_password", methods=["GET", "POST"])
def admin_reset_password():

    if not session.get("admin_otp_verified"):

        return redirect("/admin_forgot_password")

    if request.method == "POST":

        password = request.form["password"]
        confirm = request.form["confirm_password"]

        if password != confirm:

            flash(
                "Passwords do not match!",
                "danger"
            )

            return render_template(
                "admin_reset_password.html"
            )

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
        UPDATE admin
        SET password=?
        WHERE username=?
        """, (

            generate_password_hash(password),

            session["admin_reset_username"]

        ))

        conn.commit()
        conn.close()

        session.pop("admin_reset_username", None)
        session.pop("admin_reset_email", None)
        session.pop("admin_reset_otp", None)
        session.pop("admin_reset_otp_time", None)
        session.pop("admin_otp_verified", None)

        flash(
            "✅ Password Updated Successfully!",
            "success"
        )

        return redirect("/admin_login")

    return render_template(
        "admin_reset_password.html"
    )

@app.route("/verify_register_otp", methods=["GET", "POST"])
def verify_register_otp():

    if request.method == "POST":

        user_otp = request.form["otp"]

        # ===============================
        # OTP Expiry Check (5 Minutes)
        # ===============================
        if "student_otp_time" in session:

            otp_time = datetime.fromisoformat(
                session["student_otp_time"]
            )

            if datetime.now() > otp_time + timedelta(minutes=5):

                session.clear()

                flash(
                    "❌ OTP Expired! Please register again.",
                    "danger"
                )

                return redirect("/register")

        # ===============================
        # Verify OTP
        # ===============================
        if user_otp == session.get("student_otp"):

            conn = get_connection()
            cursor = conn.cursor()

            # Hash password
            hashed_password = generate_password_hash(
                session["student_password"]
            )

            cursor.execute("""
            INSERT INTO students(
            name,
            email,
            password,
            register_date
            )
            VALUES(?,?,?,?)
            """, (
                session["student_name"],
                session["student_email"],
                hashed_password,
                get_ist_string()
            ))

            conn.commit()
            conn.close()

            add_notification(
                f"🆕 New Student Registered: {session['student_name']}",
                "student"
                )

            # Clear all student session data
            session.pop("student_name", None)
            session.pop("student_email", None)
            session.pop("student_password", None)
            session.pop("student_otp", None)
            session.pop("student_otp_time", None)
            session.pop("student_resend_time", None)

            flash(
                "🎉 Registration Successful! Please Login.",
                "success"
            )

            return redirect("/login")

        else:

            flash(
                "❌ Invalid OTP!",
                "danger"
            )

    # ======================================
    # Remaining cooldown time
    # ======================================

    remaining = 0

    if "student_resend_time" in session:

        last_time = datetime.fromisoformat(
            session["student_resend_time"]
        )

        remaining = 30 - (datetime.now() - last_time).seconds

        if remaining < 0:

            remaining = 0

    return render_template(
        "verify_register_otp.html",
        remaining=remaining
    )

@app.route("/resend_register_otp")
def resend_register_otp():

    # Prevent spam (30 seconds)
    if "student_resend_time" in session:

        last_time = datetime.fromisoformat(
            session["student_resend_time"]
        )

        remaining = 30 - (datetime.now() - last_time).seconds

        if remaining > 0:

            flash(
                f"⏳ Please wait {remaining} seconds before requesting another OTP.",
                "warning"
            )

            return redirect("/verify_register_otp")

    # Generate new OTP
    otp = str(random.randint(100000,999999))

    session["student_otp"] = otp
    session["student_otp_time"] = datetime.now().isoformat()
    session["student_resend_time"] = datetime.now().isoformat()

    send_otp(
        session["student_email"],
        otp
    )

    flash(
        "✅ A new OTP has been sent.",
        "success"
    )

    return redirect("/verify_register_otp")

@app.route("/student_management")
def student_management():

    conn = get_connection()
    cursor = conn.cursor()

    # Get all students
    cursor.execute("""
        SELECT *
        FROM students
        ORDER BY id DESC
    """)

    students = cursor.fetchall()

    # Total Students
    cursor.execute("""
        SELECT COUNT(*)
        FROM students
    """)

    total_students = cursor.fetchone()[0]

    # Online Students
    cursor.execute("""
        SELECT COUNT(*)
        FROM students
        WHERE status='Online'
    """)

    online_students = cursor.fetchone()[0]

    # Registered Today
    cursor.execute("""
        SELECT COUNT(*)
        FROM students
        WHERE DATE(register_date)=DATE('now')
    """)

    today_registered = cursor.fetchone()[0]

    conn.close()

    return render_template(
        "student_management.html",
        students=students,
        total_students=total_students,
        online_students=online_students,
        today_registered=today_registered
    )

@app.route("/delete_student/<int:id>")
def delete_student(id):

    conn = get_connection()
    cursor = conn.cursor()

    # ====================================
    # Get Student Information
    # ====================================

    cursor.execute("""
    SELECT name, email
    FROM students
    WHERE id=?
    """, (id,))

    student = cursor.fetchone()

    if student is None:

        conn.close()

        return redirect("/student_management")

    student_name = student["name"]
    student_email = student["email"]

    # ====================================
    # Delete Student Feedback
    # ====================================

    cursor.execute("""
    DELETE FROM feedback
    WHERE student_name=?
    """, (student_name,))

    # ====================================
    # Delete Login History
    # ====================================

    cursor.execute("""
    DELETE FROM login_history
    WHERE email=?
    """, (student_email,))

    # ====================================
    # Delete Notifications (if linked)
    # ====================================

    cursor.execute("""
    DELETE FROM notifications
    WHERE message LIKE ?
    """, (f"%{student_name}%",))

    # ====================================
    # Delete Student Account
    # ====================================

    cursor.execute("""
    DELETE FROM students
    WHERE id=?
    """, (id,))

    conn.commit()
    conn.close()

    return redirect("/student_management?deleted=1")

@app.route("/view_student/<int:id>")
def view_student(id):

    conn = get_connection()
    cursor = conn.cursor()

    # ============================
    # Student Details
    # ============================

    cursor.execute("""
    SELECT *
    FROM students
    WHERE id=?
    """, (id,))

    student = cursor.fetchone()

    # If student not found
    if not student:

        conn.close()

        flash(
            "Student not found!",
            "danger"
        )

        return redirect("/student_management")

    # ============================
    # Total Feedback
    # ============================

    cursor.execute("""
    SELECT COUNT(*)
    FROM feedback
    WHERE student_name=?
    """, (student["name"],))

    total_feedback = cursor.fetchone()[0]

    # ============================
    # Positive Feedback
    # ============================

    cursor.execute("""
    SELECT COUNT(*)
    FROM feedback
    WHERE student_name=?
    AND sentiment='Positive'
    """, (student["name"],))

    positive = cursor.fetchone()[0]

    # ============================
    # Negative Feedback
    # ============================

    cursor.execute("""
    SELECT COUNT(*)
    FROM feedback
    WHERE student_name=?
    AND sentiment='Negative'
    """, (student["name"],))

    negative = cursor.fetchone()[0]

    # ============================
    # Neutral Feedback
    # ============================

    cursor.execute("""
    SELECT COUNT(*)
    FROM feedback
    WHERE student_name=?
    AND sentiment='Neutral'
    """, (student["name"],))

    neutral = cursor.fetchone()[0]

    # ============================
    # Recent Feedback Activity
    # ============================

    cursor.execute("""
    SELECT
        category,
        feedback,
        sentiment,
        date
    FROM feedback
    WHERE student_name=?
    ORDER BY date DESC
    LIMIT 10
    """, (student["name"],))

    recent_feedback = cursor.fetchall()

    conn.close()

    return render_template(

        "view_student.html",

        student=student,

        total_feedback=total_feedback,

        positive=positive,

        negative=negative,

        neutral=neutral,

        recent_feedback=recent_feedback

    )

def detect_intent(question):

    question = question.lower().strip()

    intents = {

        # =====================================
        # Student Count
        # =====================================

        "student_count":[

            "student",

            "students",

            "student count",

            "total students",

            "registered students",

            "number of students",

            "total users",

            "users",

            "how many students"

        ],

        # =====================================
        # Online Students
        # =====================================

        "online_students":[

            "online",

            "online students",

            "students online",

            "who is online",

            "active students",

            "logged in students"

        ],

        # =====================================
        # Offline Students
        # =====================================

        "offline_students":[

            "offline",

            "offline students",

            "logged out",

            "inactive students"

        ],

        # =====================================
        # Positive Feedback
        # =====================================

        "positive_feedback":[

            "positive",

            "positive feedback",

            "good feedback",

            "happy students",

            "positive feedbacks"

        ],

        # =====================================
        # Negative Feedback
        # =====================================

        "negative_feedback":[

            "negative",

            "negative feedback",

            "bad feedback",

            "complaints",

            "negative feedbacks"

        ],

        # =====================================
        # Neutral Feedback
        # =====================================

        "neutral_feedback":[

            "neutral",

            "neutral feedback"

        ],

        # =====================================
        # Total Feedback
        # =====================================

        "total_feedback":[

            "feedback",

            "feedback count",

            "total feedback",

            "all feedback",

            "feedback submitted"

        ],

        # =====================================
        # Top Category
        # =====================================

        "top_category":[

            "top category",

            "best category",

            "most feedback category",

            "highest feedback category",

            "category"

        ],

        # =====================================
        # Most Active Student
        # =====================================

        "top_student":[

            "top student",

            "most active student",

            "most feedback",

            "who submitted the most feedback"

        ],

        # =====================================
        # Registered Today
        # =====================================

        "registered_today":[

            "registered today",

            "today registration",

            "today registered",

            "new students today"

        ],

        # =====================================
        # Admin Count
        # =====================================

        "admin_count":[

            "admin",

            "admins",

            "admin count",

            "total admins"

        ]

    }

    for intent, keywords in intents.items():

        for keyword in keywords:

            if keyword in question:

                return intent

    return "unknown"

@app.route("/ai_assistant", methods=["GET", "POST"])
def ai_assistant():

    question = ""
    answer = ""

    if request.method == "POST":

        question = request.form["question"]

        intent = detect_intent(question)

        conn = get_connection()
        cursor = conn.cursor()

        # =====================================
        # Student Count
        # =====================================

        if intent == "student_count":

            cursor.execute("""
            SELECT COUNT(*)
            FROM students
            """)

            total = cursor.fetchone()[0]

            answer = f"""
📊 Student Statistics

Total Registered Students : {total}

✅ All registered students can log in after email verification.
"""

        # =====================================
        # Online Students
        # =====================================

        elif intent == "online_students":

            cursor.execute("""
            SELECT COUNT(*)
            FROM students
            WHERE status='Online'
            """)

            online = cursor.fetchone()[0]

            answer = f"""
🟢 Online Students

Currently Online : {online}
"""

        # =====================================
        # Offline Students
        # =====================================

        elif intent == "offline_students":

            cursor.execute("""
            SELECT COUNT(*)
            FROM students
            WHERE status='Offline'
            """)

            offline = cursor.fetchone()[0]

            answer = f"""
🔴 Offline Students

Currently Offline : {offline}
"""

        # =====================================
        # Positive Feedback
        # =====================================

        elif intent == "positive_feedback":

            cursor.execute("""
            SELECT COUNT(*)
            FROM feedback
            WHERE sentiment='Positive'
            """)

            positive = cursor.fetchone()[0]

            answer = f"""
😊 Positive Feedback Analysis

Positive Feedback : {positive}

⭐⭐⭐⭐⭐ Overall Satisfaction : Excellent

Recommendation

Continue maintaining the current teaching quality.
"""

        # =====================================
        # Negative Feedback
        # =====================================

        elif intent == "negative_feedback":

            cursor.execute("""
            SELECT COUNT(*)
            FROM feedback
            WHERE sentiment='Negative'
            """)

            negative = cursor.fetchone()[0]

            answer = f"""
⚠ Negative Feedback Analysis

Negative Feedback : {negative}

Recommendation

• Review all negative feedback.

• Improve weak areas.

• Conduct follow-up surveys.
"""

        # =====================================
        # Neutral Feedback
        # =====================================

        elif intent == "neutral_feedback":

            cursor.execute("""
            SELECT COUNT(*)
            FROM feedback
            WHERE sentiment='Neutral'
            """)

            neutral = cursor.fetchone()[0]

            answer = f"""
😐 Neutral Feedback

Neutral Feedback : {neutral}
"""

        # =====================================
        # Total Feedback
        # =====================================

        elif intent == "total_feedback":

            cursor.execute("""
            SELECT COUNT(*)
            FROM feedback
            """)

            total = cursor.fetchone()[0]

            answer = f"""
📝 Feedback Summary

Total Feedback Submitted : {total}
"""

        # =====================================
        # Top Category
        # =====================================

        elif intent == "top_category":

            cursor.execute("""
            SELECT category,
                   COUNT(*) as total
            FROM feedback
            GROUP BY category
            ORDER BY total DESC
            LIMIT 1
            """)

            row = cursor.fetchone()

            if row:

                answer = f"""
🏆 Most Discussed Category

Category : {row['category']}

Feedback Count : {row['total']}
"""

            else:

                answer = "No feedback available."

        # =====================================
        # Most Active Student
        # =====================================

        elif intent == "top_student":

            cursor.execute("""
            SELECT student_name,
                   COUNT(*) as total
            FROM feedback
            GROUP BY student_name
            ORDER BY total DESC
            LIMIT 1
            """)

            row = cursor.fetchone()

            if row:

                answer = f"""
👑 Most Active Student

Name : {row['student_name']}

Feedback Submitted : {row['total']}
"""

            else:

                answer = "No feedback available."

        # =====================================
        # Registered Today
        # =====================================

        elif intent == "registered_today":

            cursor.execute("""
            SELECT COUNT(*)
            FROM students
            WHERE DATE(register_date)=DATE('now')
            """)

            today = cursor.fetchone()[0]

            answer = f"""
🆕 Today's Registration

New Students Registered : {today}
"""

        # =====================================
        # Admin Count
        # =====================================

        elif intent == "admin_count":

            cursor.execute("""
            SELECT COUNT(*)
            FROM admin
            """)

            admins = cursor.fetchone()[0]

            answer = f"""
👨‍💼 Administrator Information

Total Administrators : {admins}
"""

        # =====================================
        # Default
        # =====================================

        else:

            answer = """
🤖 Sorry!

I couldn't understand your question.

You can ask things like:

👨 Students
• How many students?
• Student count
• Online students
• Offline students
• Registered today

📊 Feedback
• Total feedback
• Positive feedback
• Negative feedback
• Neutral feedback
• Most feedback category

👨‍🎓 Students
• Who submitted the most feedback?

👨‍💼 Admin
• Total admins
"""

        conn.close()

    return render_template(
        "ai_assistant.html",
        question=question,
        answer=answer
    )

@app.route("/notifications")
def notifications():

    if "admin" not in session:

        flash("Please login first.", "warning")
        return redirect("/admin_login")

    conn = get_connection()
    cursor = conn.cursor()

    # Get all notifications
    cursor.execute("""
    SELECT *
    FROM notifications
    ORDER BY id DESC
    """)

    notifications = cursor.fetchall()

    conn.close()

    return render_template(
        "notifications.html",
        notifications=notifications
    )

@app.route("/mark_all_read")
def mark_all_read():

    if "admin" not in session:

        return redirect("/admin_login")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    UPDATE notifications
    SET is_read = 1
    WHERE is_read = 0
    """)

    conn.commit()
    conn.close()

    return redirect("/notifications")

@app.route("/delete_notification/<int:id>")
def delete_notification(id):

    if "admin" not in session:

        return redirect("/admin_login")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    DELETE FROM notifications
    WHERE id=?
    """, (id,))

    conn.commit()
    conn.close()

    return redirect("/notifications?deleted=1")

@app.route("/delete_all_feedback")
def delete_all_feedback():

    # Check admin login
    if "admin" not in session:
        return redirect("/admin_login")

    conn = get_connection()
    cursor = conn.cursor()

    # Delete all feedback
    cursor.execute("DELETE FROM feedback")

    conn.commit()
    conn.close()

    flash(
        "All feedback deleted successfully.",
        "success"
    )

    return redirect("/dashboard_v2")

@app.route("/admin_delete_submission/<int:submission_id>")
def admin_delete_submission(submission_id):

    if "admin" not in session:
        return redirect("/admin_login")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute(
        """
        DELETE FROM feedback
        WHERE submission_id=?
        """,
        (submission_id,)
    )

    conn.commit()
    conn.close()

    flash(
        "Complete feedback submission deleted successfully.",
        "success"
    )

    return redirect("/dashboard_v2")

@app.route("/delete_all_notifications")
def delete_all_notifications():

    if "admin" not in session:

        return redirect("/admin_login")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    DELETE FROM notifications
    """)

    conn.commit()
    conn.close()

    return redirect("/notifications?cleared=1")

from datetime import datetime

@app.route("/login_history")
def login_history():

    if "admin" not in session:
        return redirect("/admin_login")

    conn = get_connection()
    cursor = conn.cursor()

    search = request.args.get("search", "").strip()

    # ===============================
    # Dashboard Cards
    # ===============================

    cursor.execute("SELECT COUNT(*) FROM login_history")
    total_logins = cursor.fetchone()[0]

    cursor.execute("""
    SELECT COUNT(*)
    FROM login_history
    WHERE status='Online'
    """)
    online_students = cursor.fetchone()[0]

    cursor.execute("""
    SELECT COUNT(*)
    FROM login_history
    WHERE status='Logged Out'
    """)
    logged_out = cursor.fetchone()[0]

    cursor.execute("""
    SELECT COUNT(*)
    FROM login_history
    WHERE DATE(login_time)=DATE('now')
    """)
    today_logins = cursor.fetchone()[0]

    # ===============================
    # Search
    # ===============================

    if search:

        cursor.execute("""
        SELECT *
        FROM login_history
        WHERE student_name LIKE ?
        OR email LIKE ?
        ORDER BY id DESC
        """, (
            f"%{search}%",
            f"%{search}%"
        ))

    else:

        cursor.execute("""
        SELECT *
        FROM login_history
        ORDER BY id DESC
        """)

    history = cursor.fetchall()

    conn.close()

    history_data = []

    for row in history:

        duration = "Still Online"

        if row["logout_time"]:

            try:

                login = datetime.strptime(
                    row["login_time"],
                    "%Y-%m-%d %H:%M:%S"
                )

                logout = datetime.strptime(
                    row["logout_time"],
                    "%Y-%m-%d %H:%M:%S"
                )

                diff = logout - login

                minutes = int(diff.total_seconds() / 60)

                if minutes < 60:

                    duration = f"{minutes} Minute(s)"

                else:

                    hours = minutes // 60
                    mins = minutes % 60

                    duration = f"{hours} Hr {mins} Min"

            except:

                duration = "-"

        history_data.append({

            "id": row["id"],

            "student_name": row["student_name"],

            "email": row["email"],

            "login_time": row["login_time"],

            "logout_time": row["logout_time"],

            "status": row["status"],

            "duration": duration

        })

    return render_template(

        "login_history.html",

        history=history_data,

        total_logins=total_logins,

        online_students=online_students,

        logged_out=logged_out,

        today_logins=today_logins,

        search=search

    )

@app.route("/export_login_history")
def export_login_history():

    if "admin" not in session:
        return redirect("/admin_login")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT *
    FROM login_history
    ORDER BY id DESC
    """)

    rows = cursor.fetchall()

    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([

        "S.No.",
        "Student Name",
        "Email",
        "Login Time",
        "Logout Time",
        "Status"

    ])

    for i, row in enumerate(rows, start=1):

        writer.writerow([

            i,
            row["student_name"],
            row["email"],
            row["login_time"],
            row["logout_time"],
            row["status"]

        ])

    output.seek(0)

    return Response(

        output.getvalue(),

        mimetype="text/csv",

        headers={

            "Content-Disposition":
            "attachment; filename=login_history.csv"

        }

    )

@app.route("/export_login_history_excel")
def export_login_history_excel():

    if "admin" not in session:
        return redirect("/admin_login")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT *
    FROM login_history
    ORDER BY id DESC
    """)

    rows = cursor.fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active

    ws.title = "Login History"

    # ==========================
    # Header
    # ==========================

    headers = [

        "S.No.",

        "Student Name",

        "Email",

        "Login Time",

        "Logout Time",

        "Status"

    ]

    header_fill = PatternFill(

        start_color="0D6EFD",

        end_color="0D6EFD",

        fill_type="solid"

    )

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(row=1, column=col)

        cell.value = header

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # ==========================
    # Data
    # ==========================

    for i, row in enumerate(rows, start=2):

        ws.cell(row=i, column=1).value = i - 1
        ws.cell(row=i, column=2).value = row["student_name"]
        ws.cell(row=i, column=3).value = row["email"]
        ws.cell(row=i, column=4).value = row["login_time"]

        logout = row["logout_time"]

        if logout is None:
            logout = "Still Online"

        ws.cell(row=i, column=5).value = logout
        ws.cell(row=i, column=6).value = row["status"]

    # Freeze header row
    ws.freeze_panes = "A2"

    # Enable filters
    ws.auto_filter.ref = ws.dimensions

    # ==========================
    # Auto Width
    # ==========================

    for column_cells in ws.columns:

        length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in column_cells
        )

        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = min(length + 5, 40)

    filename = "LoginHistory.xlsx"

    wb.save(filename)

    return send_file(
        filename,
        as_attachment=True
    )

def generate_login_history():

    conn = get_connection()
    cursor = conn.cursor()

    # ==========================
    # Statistics
    # ==========================

    cursor.execute("SELECT COUNT(*) FROM login_history")
    total = cursor.fetchone()[0]

    cursor.execute("""
    SELECT COUNT(*)
    FROM login_history
    WHERE status='Online'
    """)
    online = cursor.fetchone()[0]

    cursor.execute("""
    SELECT COUNT(*)
    FROM login_history
    WHERE status='Logged Out'
    """)
    logged_out = cursor.fetchone()[0]

    cursor.execute("""
    SELECT *
    FROM login_history
    ORDER BY id DESC
    """)

    rows = cursor.fetchall()

    conn.close()

    pdf = SimpleDocTemplate("LoginHistoryReport.pdf")

    styles = getSampleStyleSheet()

    title_style = styles["Heading1"]
    title_style.alignment = TA_CENTER

    subtitle_style = styles["Heading2"]
    subtitle_style.alignment = TA_CENTER

    elements = []

    # ==========================
    # Project Logo
    # ==========================

    logo = Image(
        "static/images/logo.png",
        width=80,
        height=80
    )

    logo.hAlign = "CENTER"

    elements.append(logo)

    elements.append(Spacer(1, 10))

    today = get_ist_time().strftime("%d %B %Y %I:%M %p")

    # ==========================
    # Heading
    # ==========================

    elements.append(
        Paragraph(
            "<font color='#0d6efd'><b>SMART FEEDBACK &amp; SUGGESTION PORTAL</b></font>",
            title_style
        )
    )

    d = Drawing(500, 1)

    d.add(Line(0, 0, 500, 0))

    elements.append(d)

    elements.append(Spacer(1, 10))

    elements.append(
        Paragraph(
            "Student Login History Report",
            subtitle_style
        )
    )

    elements.append(Spacer(1,12))

    elements.append(
        Paragraph(
            f"<b>Generated On:</b> {today}",
            styles["Normal"]
        )
    )

    elements.append(Spacer(1,15))

    # ==========================
    # Summary
    # ==========================

    elements.append(
        Paragraph(
            "<b>LOGIN SUMMARY</b>",
            styles["Heading2"]
        )
    )

    elements.append(
        Paragraph(
            f"Total Login Records : {total}",
            styles["Normal"]
        )
    )

    elements.append(
        Paragraph(
            f"Currently Online : {online}",
            styles["Normal"]
        )
    )

    elements.append(
        Paragraph(
            f"Logged Out : {logged_out}",
            styles["Normal"]
        )
    )

    elements.append(Spacer(1,18))

    # ==========================
    # Table
    # ==========================

    data = [[

        "ID",

        "Student",

        "Email",

        "Login",

        "Logout",

        "Status"

    ]]

    for i, row in enumerate(rows, start=1):

        logout = row["logout_time"]

        if logout is None:

            logout = "Still Online"

        data.append([

            i,

            row["student_name"],

            row["email"],

            row["login_time"],

            logout,

            row["status"]

        ])

    table = Table(

        data,

        colWidths=[30,80,150,110,110,70]

    )

    table.setStyle(TableStyle([

        ("BACKGROUND",(0,0),(-1,0),colors.darkblue),

        ("TEXTCOLOR",(0,0),(-1,0),colors.white),

        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ("GRID",(0,0),(-1,-1),1,colors.black),

        ("BACKGROUND",(0,1),(-1,-1),colors.beige),

        ("FONTSIZE",(0,0),(-1,-1),8),

        ("ALIGN",(0,0),(-1,-1),"CENTER"),

        ("VALIGN",(0,0),(-1,-1),"MIDDLE")

    ]))

    elements.append(table)

    elements.append(Spacer(1,20))

    elements.append(Spacer(1,20))

    line = Drawing(450,1)
    line.add(Line(0,0,450,0))
    elements.append(line)

    elements.append(Spacer(1,12))

    elements.append(Spacer(1, 20))

    line = Drawing(450, 1)
    line.add(Line(0, 0, 450, 0))
    elements.append(line)

    elements.append(Spacer(1, 10))

    footer = """
    <para align='center'>
    <b>Generated by</b><br/><br/>

    <font color='#0d6efd'><b>SMART FEEDBACK &amp; SUGGESTION PORTAL</b></font><br/><br/>

    <b>Developed By</b><br/>

    Avijit Roy<br/>

    <font color='grey'>
    This report was automatically generated by the
    Smart Feedback &amp; Suggestion Portal.
    </font>

    </para>
    """

    elements.append(
        Paragraph(
            footer,
            styles["Normal"]
        )
    )

@app.route("/pdf_login_history")
def pdf_login_history():

    if "admin" not in session:

        return redirect("/admin_login")

    generate_login_history()

    return send_file(

        "LoginHistoryReport.pdf",

        as_attachment=True

    )

@app.route("/delete_login_history/<int:id>")
def delete_login_history(id):

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    DELETE FROM login_history
    WHERE id=?
    """, (id,))

    conn.commit()
    conn.close()

    return redirect("/login_history?deleted=1")

@app.route("/clear_login_history")
def clear_login_history():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    DELETE FROM login_history
    """)

    conn.commit()
    conn.close()

    return redirect("/login_history?cleared=1")

@app.route("/clear_session")
def clear_session():

    session.clear()

    return "Session Cleared"

if __name__ == "__main__":
    app.run(debug=True)